from enum import EJECT
from scheduler import generate_timetable
from flask import Flask, render_template, request, redirect, session
import psycopg2

from werkzeug.security import generate_password_hash, check_password_hash
def get_db_connection():
    return psycopg2.connect(
        DATABASE_URL,
        sslmode="require",
        connect_timeout=10
    )
def validate_subject_feasibility(class_id, periods_per_week):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT periods_per_day, working_days FROM classes WHERE id = %s", (class_id,))
    class_data = cur.fetchone()
    conn.close()

    if not class_data:
        return True, ""

    periods_per_day, working_days = class_data
    max_weekly_slots = (periods_per_day - 1) * working_days  # minus break slot logic

    if periods_per_week > max_weekly_slots:
        return False, f"Invalid: Subject periods ({periods_per_week}) exceed weekly capacity ({max_weekly_slots})"

    return True, ""

import os

app = Flask(__name__)

# Production-safe session settings
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Strong secret key (important for login + sessions)
app.secret_key = os.environ.get("SECRET_KEY") or "super_strong_timetable_secret_2026"

DATABASE_URL = os.environ.get("DATABASE_URL")

# -------- DATABASE INIT --------
# -------- DATABASE INIT (FINAL CLEAN VERSION) --------
def init_db():
    conn = get_db_connection()
    cur = conn.cursor()

    # ADMIN TABLE (PostgreSQL version)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS admin (
        id SERIAL PRIMARY KEY,
        username TEXT UNIQUE,
        password TEXT,
        name TEXT,
        role TEXT DEFAULT 'admin'
    );
    """)

    # CLASSES TABLE
    cur.execute("""
    CREATE TABLE IF NOT EXISTS classes (
        id SERIAL PRIMARY KEY,
        name TEXT,
        periods_per_day INTEGER,
        working_days INTEGER,
        start_time TEXT,
        break_after INTEGER
    );
    """)

    # FACULTY TABLE
    cur.execute("""
    CREATE TABLE IF NOT EXISTS faculty (
        id SERIAL PRIMARY KEY,
        name TEXT,
        max_classes_per_day INTEGER
    );
    """)

    # SUBJECTS TABLE
    cur.execute("""
    CREATE TABLE IF NOT EXISTS subjects (
        id SERIAL PRIMARY KEY,
        name TEXT,
        is_lab INTEGER,
        lab_duration INTEGER,
        periods_per_week INTEGER
    );
    """)

    # SUBJECT ASSIGNMENTS TABLE
    cur.execute("""
    CREATE TABLE IF NOT EXISTS subject_assignments (
        id SERIAL PRIMARY KEY,
        subject_id INTEGER,
        class_id INTEGER,
        faculty_id INTEGER
    );
    """)

    # Insert default admin (PostgreSQL safe)
    cur.execute("SELECT * FROM admin WHERE username = %s", ("admin",))
    admin_exists = cur.fetchone()

    if not admin_exists:
        from werkzeug.security import generate_password_hash
        hashed_password = generate_password_hash("admin123")
        cur.execute(
            "INSERT INTO admin (username, password, name, role) VALUES (%s, %s, %s, %s)",
            ("admin", hashed_password, "Main Administrator", "superadmin")
        )

    conn.commit()
    cur.close()
    conn.close()
if __name__ == "__main__":
    init_db()  # ✅ only run DB init locally, not on Render import
    app.run(host="0.0.0.0", port=5000, debug=True)

# -------- LOGIN --------


@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = get_db_connection()
        cur = conn.cursor()

        # Fetch user by username ONLY
        cur.execute("SELECT username, password, name, role FROM admin WHERE username=%s", (username,))
        user = cur.fetchone()
        conn.close()

        if user:
            stored_password = user[1]

    # SMART CHECK: supports BOTH old plain passwords AND new hashed ones
            if check_password_hash(stored_password, password) or stored_password == password:
              session["user"] = user[0]
              session["admin_name"] = user[2]
              session["role"] = user[3]
              return redirect("/dashboard")

        return render_template("login.html", error="Invalid Credentials")

    return render_template("login.html")
# -------- DASHBOARD --------
@app.route("/dashboard")
def dashboard():
    if "user" not in session:
        return redirect("/")
    return render_template(
    "dashboard.html",
    user=session["user"],
    admin_name=session.get("admin_name", "Administrator")
)

# -------- ADD CLASS MODULE --------
@app.route("/add_class", methods=["GET", "POST"])
def add_class():
    if "user" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor()

    message = None

    if request.method == "POST":
        name = request.form.get("name")
        periods = request.form.get("periods")
        days = request.form.get("days")
        start_time = request.form.get("start_time")
        break_after = request.form.get("break_after")

        # Strong validation (prevents silent failures)
        if not name or not periods or not days or not break_after:
            conn.close()
            return "<h3 style='color:red;'>Missing required fields</h3><a href='/add_class'>Go Back</a>"

        # Prevent duplicate classes (VERY IMPORTANT)
        cur.execute("SELECT * FROM classes WHERE name = %s", (name,))
        existing = cur.fetchone()

        if existing:
            message = f"Class '{name}' already exists!"
        else:
            cur.execute("""
                INSERT INTO classes (name, periods_per_day, working_days, start_time, break_after)
                VALUES (%s, %s, %s, %s, %s)
            """, (name, periods, days, start_time, break_after))
            conn.commit()
            message = f"Class '{name}' added successfully!"

    # Fetch all classes to show confirmation (CRITICAL UX)
    cur.execute("SELECT * FROM classes")
    classes = cur.fetchall()

    conn.close()
    return render_template("manage_data.html", message=message, classes=classes)
# -------- ADD ASSIGNMENT MODULE (CRITICAL FOR TIMETABLE) --------
@app.route("/add_assignment", methods=["POST"])
def add_assignment():
    if "user" not in session:
        return redirect("/")

    # DEBUG (do not remove until stable)
    print("ASSIGNMENT FORM DATA:", request.form)

    subject_id = request.form.get("subject_id")
    class_id = request.form.get("class_id")
    faculty_id = request.form.get("faculty_id")

    # HARD VALIDATION (prevents your recurring crashes)
    if not subject_id or not class_id or not faculty_id:
        return "<h3 style='color:red;'>Missing assignment fields (Subject/Class/Faculty)</h3><a href='/manage_data'>Go Back</a>"

    conn = get_db_connection()
    cur = conn.cursor()

    # Prevent duplicate assignment (professional logic)
    cur.execute("""
        SELECT * FROM subject_assignments
        WHERE subject_id = %s AND class_id = %s AND faculty_id = %s
    """, (subject_id, class_id, faculty_id))

    exists = cur.fetchone()

    if exists:
        conn.close()
        return "<h3 style='color:orange;'>This assignment already exists.</h3><a href='/manage_data'>Go Back</a>"

    # Insert assignment
    cur.execute("""
        INSERT INTO subject_assignments (subject_id, class_id, faculty_id)
        VALUES (%s, %s, %s)
    """, (subject_id, class_id, faculty_id))

    conn.commit()
    conn.close()

    # IMPORTANT: redirect back to manage panel (your UI structure)
    return redirect("/manage_data")
# -------- ADD FACULTY MODULE --------
# -------- ADD FACULTY MODULE --------
# -------- ADD FACULTY MODULE --------
@app.route("/add_faculty", methods=["GET", "POST"])
def add_faculty():
    if "user" not in session:
        return redirect("/")

    if request.method == "POST":
        name = request.form.get("name")
        max_daily = request.form.get("max_classes")
        print("form data:",request.form)

        # Strict validation
        if not name or not max_daily:
            return "<h3 style='color:red;'>Form Error: Missing Faculty Name or Max Daily Field</h3><a href='/add_faculty'>Go Back</a>"

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO faculty (name, max_classes_per_day)
            VALUES (%s, %s)
        """, (name, max_daily))
        conn.commit()
        conn.close()

        # Redirect ONLY after successful POST
        return redirect("/manage_data")

    # THIS must be outside POST block (VERY IMPORTANT)
    return render_template("add_faculty.html")

# -------- ADD SUBJECT MODULE --------
@app.route("/add_subject", methods=["GET", "POST"])
def add_subject():
    if "user" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        # Debug (keep for now)
        print("SUBJECT FORM DATA:", request.form)

        # Safe extraction (NO crashes)
        name = request.form.get("name")
        periods_per_week = request.form.get("periods") or request.form.get("periods_per_week")
        is_lab = 1 if request.form.get("is_lab") else 0
        lab_duration = request.form.get("lab_duration")

        # Validation
        if not name or not periods_per_week:
            conn.close()
            return "<h3 style='color:red;'>Missing required fields</h3><a href='/manage_data'>Go Back</a>"

        # Convert values safely
        try:
            periods_per_week = int(periods_per_week)
        except:
            conn.close()
            return "<h3 style='color:red;'>Invalid periods value</h3><a href='/manage_data'>Go Back</a>"

        # Lab logic (VERY IMPORTANT)
        if is_lab:
            lab_duration = int(lab_duration) if lab_duration else 2
        else:
            lab_duration = 1  # Theory always single slot

        # ONLY insert into subjects table (NO faculty here)
        cur.execute("""
            INSERT INTO subjects (name, is_lab, lab_duration, periods_per_week)
            VALUES (%s, %s, %s, %s)
        """, (name.strip(), is_lab, lab_duration, periods_per_week))

        conn.commit()
        conn.close()

        return redirect("/manage_data")

    conn.close()
    return redirect("/manage_data")
# -------- GENERATE TIMETABLE --------
@app.route("/generate")
def generate():
    if "user" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor()

    # ===== 1. FETCH CLASSES (CORRECT) =====
    cur.execute("SELECT * FROM classes")
    class_rows = cur.fetchall()
    print("TOTAL CLASSES IN DB:", len(class_rows))
    print("CLASS ROWS:", class_rows)

    classes = []
    for c in class_rows:
        classes.append({
            "id": c[0],
            "name": c[1],
            "periods_per_day": c[2],
            "working_days": c[3],
            "start_time": c[4],
            "break_after": c[5]
        })

    # ===== 2. FETCH SUBJECT ASSIGNMENTS =====
    cur.execute("SELECT * FROM subject_assignments")
    assignments_raw = cur.fetchall()

    # ===== 3. FETCH MASTER SUBJECTS =====
    cur.execute("SELECT * FROM subjects")
    subjects_raw = cur.fetchall()

    # ===== 4. BUILD SUBJECT LIST (CORRECT INDEXING) =====
    subjects = []

    for a in assignments_raw:
        subject_id = a[1]
        class_id = a[2]
        faculty_id = a[3]

        for s in subjects_raw:
            if s[0] == subject_id:
                subjects.append({
                    "name": s[1],
                    "class_id": class_id,
                    "faculty_id": faculty_id,
                    # IMPORTANT: Correct indexes based on your DB structure
                    "is_lab": bool(s[2]),
                    "lab_duration": s[3] if s[3] else 1,
                    "periods_per_week": s[4]
                })

    # ===== 5. FETCH FACULTY LIMITS =====
    cur.execute("SELECT id, max_classes_per_day FROM faculty")
    faculty_rows = cur.fetchall()
    faculty_limits = {f[0]: f[1] for f in faculty_rows}

    conn.close()

    # DEBUG (remove later)
    print("CLASSES:", classes)
    print("SUBJECTS:", subjects)

    # ===== 6. GENERATE TIMETABLE =====
    timetable = generate_timetable(classes, subjects, faculty_limits)

    # 🔥 STORE timetable globally for export
    session["last_timetable"] = timetable

    return render_template("view_timetable.html", timetable=timetable)
@app.route("/manage_data")
def manage_data():
    if "user" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM classes")
    classes = cur.fetchall()

    cur.execute("SELECT * FROM faculty")
    faculty = cur.fetchall()

    cur.execute("SELECT * FROM subjects")
    subjects = cur.fetchall()

    # 🔥 ADD THIS (Missing critical table)
    cur.execute("""
        SELECT sa.id, s.name, sa.class_id, sa.faculty_id
        FROM subject_assignments sa
        JOIN subjects s ON sa.subject_id = s.id
    """)
    assignments = cur.fetchall()

    conn.close()

    return render_template(
        "manage_data.html",
        classes=classes,
        faculty=faculty,
        subjects=subjects,
        assignments=assignments
    )
# ===== DELETE FACULTY =====
@app.route("/delete_faculty/<int:faculty_id>")
def delete_faculty(faculty_id):
    if "user" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor()

    # Delete related assignments first (safe delete)
    cur.execute("DELETE FROM subject_assignments WHERE faculty_id = %s", (faculty_id,))
    cur.execute("DELETE FROM faculty WHERE id = %s", (faculty_id,))

    conn.commit()
    conn.close()
    # Invalidate old timetable (CRITICAL)
    session.pop("last_timetable", None)

    return redirect("/manage_data")

# ===== DELETE CLASS =====
@app.route("/delete_class/<int:class_id>")
def delete_class(class_id):
    if "user" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor()

    # Remove assignments linked to this class
    cur.execute("DELETE FROM subject_assignments WHERE class_id = %s", (class_id,))
    cur.execute("DELETE FROM classes WHERE id = %s", (class_id,))

    conn.commit()
    conn.close()
    return redirect("/manage_data")


# ===== DELETE SUBJECT =====
@app.route("/delete_subject/<int:subject_id>")
def delete_subject(subject_id):
    if "user" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor()

    # Delete assignments first (important for foreign logic)
    cur.execute("DELETE FROM subject_assignments WHERE subject_id = %s", (subject_id,))
    cur.execute("DELETE FROM subjects WHERE id = %s", (subject_id,))

    conn.commit()
    conn.close()
    return redirect("/manage_data")


# ===== DELETE SUBJECT ASSIGNMENT (MOST CRITICAL) =====
@app.route("/delete_assignment/<int:assignment_id>")
def delete_assignment(assignment_id):
    if "user" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM subject_assignments WHERE id = %s", (assignment_id,))

    conn.commit()
    conn.close()
    return redirect("/manage_data")
# -------- LOGOUT --------
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

from flask import send_file
import openpyxl

from flask import send_file
import openpyxl

# ===== FULL DATABASE RESET (ADMIN ONLY) =====
@app.route("/reset_database", methods=["POST"])
def reset_database():
    if "user" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor()

    # Delete all core data (SAFE ORDER - avoids foreign issues)
    cur.execute("DELETE FROM subject_assignments")
    cur.execute("DELETE FROM subjects")
    cur.execute("DELETE FROM faculty")
    cur.execute("DELETE FROM classes")

    # Reset auto increment counters (clean IDs)
    cur.execute("TRUNCATE subject_assignments RESTART IDENTITY CASCADE")
    cur.execute("TRUNCATE subjects RESTART IDENTITY CASCADE")
    cur.execute("TRUNCATE faculty RESTART IDENTITY CASCADE")
    cur.execute("TRUNCATE classes RESTART IDENTITY CASCADE")

    # Keep admin table (DO NOT DELETE ADMINS)
    # But clear timetable session
    conn.commit()
    conn.close()

    # Clear generated timetable cache
    session.pop("last_timetable", None)

    return "<h2 style='color:red;'>Database Reset Successful</h2><a href='/dashboard'>Back to Dashboard</a>"
@app.route("/export_excel")
def export_excel():
    if "user" not in session:
        return redirect("/")

    # 🔥 Use the EXACT timetable shown on UI
    timetable = session.get("last_timetable")

    if not timetable:
        return "No timetable generated yet. Please generate timetable first."

    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for class_name, table in timetable.items():
        ws = wb.create_sheet(title=class_name)

        # Header
        ws.cell(row=1, column=1, value="Day")
        periods_count = len(next(iter(table.values())))
        for p in range(periods_count):
            ws.cell(row=1, column=p + 2, value=f"Period {p + 1}")

        # Fill timetable data
        row_num = 2
        for day, periods in table.items():
            ws.cell(row=row_num, column=1, value=day)
            for col, subject in enumerate(periods, start=2):
                ws.cell(row=row_num, column=col, value=subject)
            row_num += 1

    file_path = "timetable.xlsx"
    wb.save(file_path)

    return send_file(file_path, as_attachment=True)
    
@app.route("/faculty_timetable")
def faculty_timetable():
    if "user" not in session:
        return redirect("/")

    timetable = session.get("last_timetable")
    if not timetable:
        return "Generate timetable first."

    conn = get_db_connection()
    cur = conn.cursor()

    # Get all faculty
    cur.execute("SELECT id, name FROM faculty")
    faculty_rows = cur.fetchall()

    # Get class name -> id mapping
    cur.execute("SELECT id, name FROM classes")
    class_rows = cur.fetchall()
    class_name_to_id = {c[1]: c[0] for c in class_rows}

    # Get subject id -> name mapping
    cur.execute("SELECT id, name FROM subjects")
    subject_rows = cur.fetchall()
    subject_id_to_name = {s[0]: s[1] for s in subject_rows}

    # Get TRUE subject assignments (CRITICAL TABLE)
    cur.execute("SELECT subject_id, class_id, faculty_id FROM subject_assignments")
    assignment_rows = cur.fetchall()

    conn.close()

    # Build mapping: (class_id, subject_name) -> faculty_id
    subject_faculty_map = {}
    for subject_id, class_id, faculty_id in assignment_rows:
        subject_name = subject_id_to_name.get(subject_id)
        if subject_name:
            subject_faculty_map[(class_id, subject_name)] = faculty_id

    # Initialize faculty timetable structure
    faculty_tables = {}
    for fid, fname in faculty_rows:
        faculty_tables[fname] = {}

    # Populate faculty timetable using REAL mapping
    for class_name, class_table in timetable.items():
        class_id = class_name_to_id.get(class_name)
        if not class_id:
            continue

        for day, periods in class_table.items():
            for slot_index, subject in enumerate(periods):

                # Skip non-teaching slots
                if subject in ["Free", "BREAK"]:
                    continue

                # Remove "(Lab)" suffix for matching with DB
                clean_subject = subject.replace(" (Lab)", "")

                # Get correct faculty from DB mapping
                faculty_id = subject_faculty_map.get((class_id, clean_subject))
                if not faculty_id:
                    continue

                # Get faculty name
                faculty_name = next((f[1] for f in faculty_rows if f[0] == faculty_id), None)
                if not faculty_name:
                    continue

                # Initialize day row if not exists
                if day not in faculty_tables[faculty_name]:
                    faculty_tables[faculty_name][day] = ["Free"] * len(periods)

                # Assign correct class-subject entry
                faculty_tables[faculty_name][day][slot_index] = f"{class_name} - {subject}"

    return render_template("faculty_timetable.html", faculty_tables=faculty_tables)

@app.route("/conflict_report")
def conflict_report():
    if "user" not in session:
        return redirect("/")

    timetable = session.get("last_timetable")
    if not timetable:
        return "Generate timetable first."

    import psycopg2
    conn = get_db_connection()
    cur = conn.cursor()

    # ===== FETCH DATABASE MASTER DATA =====
    cur.execute("SELECT id, name, max_classes_per_day FROM faculty")
    faculty_rows = cur.fetchall()

    cur.execute("SELECT id, name FROM classes")
    class_rows = cur.fetchall()

    cur.execute("SELECT id, name FROM subjects")
    subject_rows = cur.fetchall()

    cur.execute("SELECT subject_id, class_id, faculty_id FROM subject_assignments")
    assignment_rows = cur.fetchall()

    conn.close()

    # ===== SAFE NORMALIZED MAPPINGS (CRITICAL) =====
    class_name_to_id = {c[1]: c[0] for c in class_rows}
    faculty_id_to_name = {f[0]: f[1] for f in faculty_rows}
    faculty_limits = {f[1]: (f[2] if f[2] else 4) for f in faculty_rows}

    subject_id_to_name = {}
    for sid, sname in subject_rows:
        if sname:
            subject_id_to_name[sid] = sname.strip().lower()

    # (class_id, subject_name) → faculty_name
    assignment_map = {}
    for subject_id, class_id, faculty_id in assignment_rows:
        subject_name = subject_id_to_name.get(subject_id)
        faculty_name = faculty_id_to_name.get(faculty_id)
        if subject_name and faculty_name:
            assignment_map[(class_id, subject_name)] = faculty_name

    # ===== ANALYTICS STRUCTURES =====
    faculty_daily_load = {f[1]: {} for f in faculty_rows}
    faculty_total_load = {f[1]: 0 for f in faculty_rows}

    true_overload_days = []
    optimal_days = []
    balanced_days = []
    underutilized_days = []

    # ===== MAIN TIMETABLE ANALYSIS =====
    for class_name, class_table in timetable.items():
        class_id = class_name_to_id.get(class_name)
        if not class_id:
            continue

        for day, periods in class_table.items():
            total_slots = len(periods)

            teaching = 0
            free = 0
            break_count = 0

            for subject in periods:
                if subject == "BREAK":
                    break_count += 1
                    continue

                if subject == "Free":
                    free += 1
                    continue

                teaching += 1

                # Normalize subject for correct faculty mapping
                clean_subject = subject.replace(" (Lab)", "").strip().lower()
                faculty_name = assignment_map.get((class_id, clean_subject))

                if faculty_name:
                    faculty_total_load[faculty_name] += 1
                    faculty_daily_load[faculty_name][day] = (
                        faculty_daily_load[faculty_name].get(day, 0) + 1
                    )

            # ===== BREAK-AWARE ACADEMIC LOAD LOGIC =====
            usable_slots = total_slots - break_count  # real teaching capacity

            # True academic overload (should almost never happen)
            if teaching > usable_slots:
                true_overload_days.append(
                    f"{class_name} is logically overloaded on {day} ({teaching}/{usable_slots} teaching slots)"
                )

            # Optimal dense schedule (BEST institutional pattern)
            elif teaching == usable_slots:
                optimal_days.append(
                    f"{class_name} is optimally scheduled on {day} ({teaching}/{usable_slots} teaching slots)"
                )

            # Well balanced academic distribution
            elif teaching >= usable_slots - 1:
                balanced_days.append(
                    f"{class_name} is well balanced on {day} ({teaching} teaching, {free} free)"
                )

            # Underutilized timetable (too many free periods)
            else:
                underutilized_days.append(
                    f"{class_name} is underutilized on {day} ({teaching} teaching, {free} free)"
                )

    # ===== FACULTY ANALYSIS (REAL CONSTRAINT CHECK) =====
    underutilized_faculty = []
    faculty_limit_violations = []

    for fname in faculty_total_load:
        total = faculty_total_load[fname]
        daily_limit = faculty_limits.get(fname, 4)

        # Unused faculty detection
        if total == 0:
            underutilized_faculty.append(
                f"{fname} has 0 assigned periods (Unused Faculty Resource)"
            )

        # Daily workload violation check
        for day, load in faculty_daily_load[fname].items():
            if load > daily_limit:
                faculty_limit_violations.append(
                    f"{fname} exceeds daily limit on {day} ({load}/{daily_limit})"
                )

    # ===== FINAL REPORT OBJECT =====
    report = {
        "faculty_overload": faculty_limit_violations,
        "underutilized_faculty": underutilized_faculty,
        "true_overload_days": true_overload_days,
        "optimal_days": optimal_days,
        "balanced_days": balanced_days,
        "underutilized_days": underutilized_days,
        "faculty_total_load": faculty_total_load
    }

    return render_template("conflict_report.html", report=report)
@app.route("/edit_faculty/<int:faculty_id>", methods=["GET", "POST"])
def edit_faculty(faculty_id):
    if "user" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        name = request.form["name"]
        max_classes = request.form["max_classes"]

        cur.execute("""
            UPDATE faculty
            SET name = %s, max_classes_per_day = %s
            WHERE id = %s
        """, (name, max_classes, faculty_id))

        conn.commit()
        conn.close()
        return redirect("/manage_data")

    # GET request (load existing data)
    cur.execute("SELECT * FROM faculty WHERE id = %s", (faculty_id,))
    faculty = cur.fetchone()
    conn.close()

    return render_template("edit_faculty.html", faculty=faculty)
@app.route("/edit_subject/<int:subject_id>", methods=["GET", "POST"])
def edit_subject(subject_id):
    if "user" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        name = request.form["name"]
        periods = request.form["periods_per_week"]
        is_lab = 1 if request.form.get("is_lab") else 0
        lab_duration = request.form["lab_duration"]

        cur.execute("""
            UPDATE subjects
            SET name = %s, periods_per_week = %s, is_lab = %s, lab_duration = %s
            WHERE id = %s
        """, (name, periods, is_lab, lab_duration, subject_id))

        conn.commit()
        conn.close()
        return redirect("/manage_data")

    # GET (load existing subject)
    cur.execute("SELECT * FROM subjects WHERE id = %s", (subject_id,))
    subject = cur.fetchone()  # IMPORTANT: variable name = subject
    conn.close()

    return render_template("edit_subject.html", subject=subject)
@app.route("/edit_class/<int:class_id>", methods=["GET", "POST"])
def edit_class(class_id):
    if "user" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        name = request.form["name"]
        periods = request.form["periods_per_day"]
        days = request.form["working_days"]
        start_time = request.form["start_time"]
        break_after = request.form["break_after"]

        cur.execute("""
            UPDATE classes
            SET name = %s, periods_per_day = %s, working_days = %s, 
                start_time = %s, break_after = %s
            WHERE id = %s
        """, (name, periods, days, start_time, break_after, class_id))

        conn.commit()
        conn.close()
        return redirect("/manage_data")

    cur.execute("SELECT * FROM classes WHERE id = %s", (class_id,))
    class_data = cur.fetchone()  # MUST match template variable name
    conn.close()

    return render_template("edit_class.html", class_data=class_data)
@app.route("/update_faculty_inline", methods=["POST"])
def update_faculty_inline():
    if "user" not in session:
        return redirect("/")

    fid = request.form["id"]
    name = request.form["name"]
    max_classes = request.form["max_classes"]

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE faculty
        SET name = %s, max_classes_per_day = %s
        WHERE id = %s
    """, (name, max_classes, fid))

    conn.commit()
    conn.close()

    session.pop("last_timetable", None)  # invalidate old timetable
    return redirect("/manage_data")
@app.route("/update_subject_inline", methods=["POST"])
def update_subject_inline():
    if "user" not in session:
        return redirect("/")

    sid = request.form["id"]
    name = request.form["name"]
    periods = int(request.form["periods"])
    is_lab = 1 if request.form.get("is_lab") else 0
    lab_duration = int(request.form.get("lab_duration") or (3 if request.form.get("is_lab") else 1))

    conn = get_db_connection()
    cur = conn.cursor()

    # Get all classes where this subject is assigned
    cur.execute("SELECT class_id FROM subject_assignments WHERE subject_id = %s", (sid,))
    assigned_classes = cur.fetchall()

    # VALIDATION CHECK
    for (class_id,) in assigned_classes:
        valid, message = validate_subject_feasibility(class_id, periods)
        if not valid:
            conn.close()
            return f"<h2 style='color:red;'>Validation Error: {message}</h2><a href='/manage_data'>Go Back</a>"

    # Lab logical validation
    if is_lab and lab_duration < 2:
        conn.close()
        return "<h2 style='color:red;'>Validation Error: Lab duration must be at least 2 periods.</h2><a href='/manage_data'>Go Back</a>"

    cur.execute("""
        UPDATE subjects
        SET name = %s, periods_per_week = %s, is_lab = %s, lab_duration = %s
        WHERE id = %s
    """, (name, periods, is_lab, lab_duration, sid))

    conn.commit()
    conn.close()

    session.pop("last_timetable", None)
    return redirect("/manage_data")
@app.route("/update_class_inline", methods=["POST"])
def update_class_inline():
    if "user" not in session:
        return redirect("/")

    cid = request.form["id"]
    name = request.form["name"]
    periods = int(request.form["periods"])
    days = int(request.form["days"])
    start_time = request.form["start_time"]
    break_after = int(request.form["break_after"])

    # HARD LOGICAL VALIDATIONS
    if break_after >= periods:
        return "<h2 style='color:red;'>Invalid: Break period must be LESS than total periods.</h2><a href='/manage_data'>Go Back</a>"

    if periods < 4:
        return "<h2 style='color:red;'>Invalid: Periods per day too low for college timetable.</h2><a href='/manage_data'>Go Back</a>"

    if days < 1 or days > 7:
        return "<h2 style='color:red;'>Invalid: Working days must be between 1 and 7.</h2><a href='/manage_data'>Go Back</a>"

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE classes
        SET name=%s, periods_per_day=%s, working_days=%s, start_time=%s, break_after=%s
        WHERE id=%s
    """, (name, periods, days, start_time, break_after, cid))

    conn.commit()
    conn.close()

    session.pop("last_timetable", None)
    return redirect("/manage_data")
@app.route("/analytics")
def analytics():
    if "user" not in session:
        return redirect("/")

    timetable = session.get("last_timetable")
    if not timetable:
        return "<h2>No timetable generated yet. Generate timetable first.</h2>"

    conn = get_db_connection()
    cur = conn.cursor()

    # Get faculty mapping
    cur.execute("SELECT id, name FROM faculty")
    faculty_rows = cur.fetchall()

    cur.execute("SELECT id, name FROM subjects")
    subject_rows = cur.fetchall()

    cur.execute("SELECT subject_id, class_id, faculty_id FROM subject_assignments")
    assignments = cur.fetchall()

    cur.execute("SELECT id, name FROM classes")
    classes = cur.fetchall()

    conn.close()

    faculty_map = {f[0]: f[1] for f in faculty_rows}
    subject_map = {s[0]: s[1].strip().lower() for s in subject_rows}
    class_map = {c[1]: c[0] for c in classes}

    assignment_lookup = {}
    for sub_id, class_id, fac_id in assignments:
        sub_name = subject_map.get(sub_id)
        if sub_name:
            assignment_lookup[(class_id, sub_name)] = faculty_map.get(fac_id, "Unknown")

    # Calculate faculty load
    faculty_load = {f[1]: 0 for f in faculty_rows}

    for class_name, table in timetable.items():
        class_id = class_map.get(class_name)
        if not class_id:
            continue

        for day, periods in table.items():
            for subject in periods:
                if subject in ["Free", "BREAK"]:
                    continue

                clean = subject.replace(" (Lab)", "").strip().lower()
                faculty = assignment_lookup.get((class_id, clean))

                if faculty:
                    faculty_load[faculty] += 1

    # Prepare chart data
    labels = list(faculty_load.keys())
    values = list(faculty_load.values())

    return render_template("analytics.html", labels=labels, values=values)

# -------- ADMIN MANAGEMENT --------
@app.route("/manage_admin", methods=["GET", "POST"])
def manage_admin():
    if "user" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        new_username = request.form["username"]
        new_password = request.form["password"]
        new_name = request.form["name"]

        # Prevent duplicate usernames
        cur.execute("SELECT * FROM admin WHERE username = %s", (new_username,))
        exists = cur.fetchone()

        if exists:
            conn.close()
            return "<h3 style='color:red;'>Admin username already exists!</h3><a href='/manage_admin'>Go Back</a>"

        hashed_password = generate_password_hash(new_password)

        cur.execute(
            "INSERT INTO admin (username, password, name) VALUES (%s, %s, %s)",
             (new_username, hashed_password, new_name)
        )
        conn.commit()
        conn.close()
        return redirect("/manage_admin")

    # Fetch all admins
    cur.execute("SELECT id, username, name FROM admin")
    admins = cur.fetchall()
    conn.close()

    return render_template("manage_admin.html", admins=admins)

@app.route("/delete_admin/<int:admin_id>")
def delete_admin(admin_id):
    if "user" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor()

    # Prevent deleting last admin
    cur.execute("SELECT COUNT(*) FROM admin")
    count = cur.fetchone()[0]

    if count <= 1:
        conn.close()
        return "<h3 style='color:red;'>Cannot delete the last admin!</h3><a href='/manage_admin'>Go Back</a>"

    cur.execute("DELETE FROM admin WHERE id = %s", (admin_id,))
    conn.commit()
    conn.close()

    return redirect("/manage_admin")

@app.route("/edit_profile", methods=["GET", "POST"])
def edit_profile():
    if "user" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        new_name = request.form["name"]
        new_password = request.form["password"]

        if new_password:
            hashed = generate_password_hash(new_password)
            cur.execute(
                "UPDATE admin SET name=%s, password=%s WHERE username=%s",
                (new_name, hashed, session["user"])
            )
        else:
            cur.execute(
                "UPDATE admin SET name=%s WHERE username=%s",
                (new_name, session["user"])
            )

        conn.commit()
        conn.close()

        session["admin_name"] = new_name
        return redirect("/dashboard")

    cur.execute("SELECT name FROM admin WHERE username=%s", (session["user"],))
    admin = cur.fetchone()
    conn.close()

    return render_template("edit_profile.html", admin=admin)

@app.route("/db_test")
def db_test():
    try:
        conn = psycopg2.connect(os.getenv("DATABASE_URL"))
        cur = conn.cursor()
        cur.execute("SELECT version();")
        version = cur.fetchone()
        conn.close()
        return f"Database Connected: {version}"
    except Exception as e:
        return f"Database Error: {e}"
    
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)